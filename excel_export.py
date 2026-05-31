"""
excel_export.py
===============
Gera o arquivo Excel em memória (bytes) para download via Streamlit.
Reutiliza a lógica de formatação do quantai_export.py,
adaptada para retornar bytes em vez de salvar em disco.

3 abas:
    1. Quantitativos  — tabela completa com fórmulas
    2. Resumo         — agrupamento por categoria
    3. Critérios      — critérios de medição, mapeamento IFC e normas
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

C_BLUE_DARK  = "00427A"
C_BLUE_MED   = "185FA5"
C_BLUE_LIGHT = "D6E8F7"
C_TEAL       = "0F6E56"
C_TEAL_LIGHT = "D4EDE6"
C_GRAY_HEAD  = "F1F0EB"
C_WHITE      = "FFFFFF"
C_YELLOW     = "FFF3CD"


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


CRITERIOS = [
    ("Revestimento — piso",   "IfcCovering (FLOORING)",      "m²",  "Área bruta projetada em planta, descontando aberturas > 0,25 m². Parâmetro: NetArea."),
    ("Revestimento — parede", "IfcCovering (CLADDING)",      "m²",  "Área líquida de face; vãos descontados automaticamente. Parâmetro: GrossArea."),
    ("Forro / teto",          "IfcCovering (CEILING)",       "m²",  "Área projetada em planta, mesmo critério do piso."),
    ("Rodapé",                "IfcCovering (BASEBOARD)",     "m",   "Perímetro interno do ambiente, descontando vãos de portas > 0,60 m. Parâmetro: Length."),
    ("Portas",                "IfcDoor",                     "un",  "Contagem por folha. Portas duplas = 2 un."),
    ("Janelas",               "IfcWindow",                   "un",  "Contagem por módulo. Área de vão como coluna auxiliar."),
    ("Louças sanitárias",     "IfcSanitaryTerminal",         "un",  "Contagem por aparelho, classificada por PredefinedType."),
    ("Impermeabilização",     "IfcSlab (ROOF)",              "m²",  "Área de laje + 10% para raios e emendas."),
]

NORMAS = [
    ("SINAPI / CEF", "Unidades e critérios de medição — referência nacional para orçamentos públicos e privados."),
    ("TCPO — PINI",  "Tabelas de Composições de Preços: critérios de desconto de vãos e arredondamentos."),
    ("NBR 12721",    "Avaliação de custos unitários — define áreas computáveis e não computáveis."),
    ("ISO 19650",    "Organização da informação BIM — nomenclatura de entidades e parâmetros IFC."),
]

IFC_MAP = [
    ("IfcCovering",         "FLOORING / CLADDING / CEILING / BASEBOARD", "Revestimentos de piso, parede, teto e rodapé."),
    ("IfcDoor",             "—",                                          "Portas. Parâmetros: OverallWidth, OverallHeight."),
    ("IfcWindow",           "—",                                          "Janelas. Parâmetros: OverallWidth, OverallHeight."),
    ("IfcSanitaryTerminal", "SINK / WC / SHOWER / BATH",                 "Louças sanitárias, diferenciadas por PredefinedType."),
    ("IfcSpace",            "—",                                          "Ambientes — associam itens ao cômodo e pavimento."),
    ("IfcBuildingStorey",   "—",                                          "Pavimentos — filtram a extração por andar."),
    ("IfcSlab",             "ROOF",                                       "Lajes de cobertura para impermeabilização."),
]


def _aba_quantitativos(wb, rows, estado, regime):
    ws = wb.active
    ws.title = "Quantitativos"

    ws.merge_cells("A1:I1")
    ws["A1"] = "QuantAI — Quantitativos de Arquitetura"
    ws["A1"].font = _font(bold=True, color=C_WHITE, size=13)
    ws["A1"].fill = _fill(C_BLUE_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:I2")
    regime_label = "Desonerado" if "desone" in regime.lower() else "Não desonerado"
    ws["A2"] = (
        f"Referência SINAPI · Estado: {estado.upper()} · "
        f"Regime: {regime_label} · Data: {date.today().strftime('%d/%m/%Y')}"
    )
    ws["A2"].font = _font(italic=True, color=C_WHITE, size=9)
    ws["A2"].fill = _fill(C_BLUE_MED)
    ws["A2"].alignment = _align("center")

    ws.merge_cells("A3:I3")
    ws["A3"] = (
        "M7T2 — Zigurat Institute of Technology · Prof. Thomas Takeuchi · Aluna: Renata Rocha · "
        "Preços: API SINAPI do Orçamentador (orcamentador.com.br/api)"
    )
    ws["A3"].font = _font(italic=True, color="555555", size=8)
    ws["A3"].fill = _fill(C_YELLOW)
    ws["A3"].alignment = _align("center")
    ws.row_dimensions[3].height = 14

    headers = ["Cód. SINAPI", "Elemento (categoria)", "Entidade IFC",
               "Pavimento", "Categoria", "Unidade",
               "Quantidade", "Preço unit. (R$)", "Total (R$)"]
    widths   = [12, 38, 20, 12, 14, 9, 11, 16, 16]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font = _font(bold=True, color=C_WHITE, size=9)
        cell.fill = _fill(C_BLUE_DARK)
        cell.alignment = _align("center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[5].height = 18
    ws.freeze_panes = "A6"

    rn = 6
    for i, row in enumerate(rows):
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
        pu = row.get("preco_unit")
        vals = [
            row.get("cod", ""),
            row.get("item", ""),
            row.get("ifc", ""),
            row.get("pav", ""),
            row.get("cat", ""),
            row.get("un", "").replace("m2", "m²"),
            row.get("qtd"),
            pu,
            f"=G{rn}*H{rn}" if pu else None,
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=rn, column=ci, value=val)
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.font = _font(size=9)
            if ci in (7, 8, 9):
                cell.alignment = _align("right")
                if ci in (8, 9) and val is not None:
                    cell.number_format = 'R$ #,##0.00'
                elif ci == 7:
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = _align("left")
        rn += 1

    total_row = rn
    ws.merge_cells(f"A{total_row}:F{total_row}")
    for col in range(1, 10):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = _fill(C_TEAL)
        cell.font = _font(bold=True, color=C_WHITE, size=9)
        cell.border = _border()

    ws[f"A{total_row}"].value = "TOTAL ESTIMADO"
    ws[f"A{total_row}"].alignment = _align("right")

    ws[f"I{total_row}"].value = f"=SUM(I6:I{total_row-1})"
    ws[f"I{total_row}"].number_format = 'R$ #,##0.00'
    ws[f"I{total_row}"].alignment = _align("right")

    ws[f"G{total_row}"].value = f"=SUM(G6:G{total_row-1})"
    ws[f"G{total_row}"].number_format = '#,##0.00'
    ws[f"G{total_row}"].alignment = _align("right")


def _aba_resumo(wb, rows):
    ws = wb.create_sheet("Resumo por Categoria")

    ws.merge_cells("A1:E1")
    ws["A1"] = "Resumo por Categoria"
    ws["A1"].font = _font(bold=True, color=C_WHITE, size=12)
    ws["A1"].fill = _fill(C_BLUE_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 20

    for ci, (h, w) in enumerate(zip(
        ["Categoria", "Itens", "Total quantidade*", "Total (R$)", "% do custo"],
        [22, 8, 18, 18, 12]
    ), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = _font(bold=True, color=C_WHITE, size=9)
        cell.fill = _fill(C_BLUE_MED)
        cell.alignment = _align("center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    cats: dict[str, dict] = {}
    for row in rows:
        c = row.get("cat", "Outros")
        if c not in cats:
            cats[c] = {"n": 0, "total": 0.0}
        cats[c]["n"] += 1
        t = row.get("total")
        if t:
            cats[c]["total"] += float(t)

    grand = sum(v["total"] for v in cats.values()) or 1
    rn = 4
    for i, (cat, v) in enumerate(cats.items()):
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
        pct = v["total"] / grand * 100
        for ci, val in enumerate([cat, v["n"], "—", v["total"] or "—", pct], 1):
            cell = ws.cell(row=rn, column=ci, value=val)
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.font = _font(size=9)
            if ci == 4 and isinstance(val, float):
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = _align("right")
            elif ci == 5:
                cell.number_format = '0.0"%"'
                cell.alignment = _align("right")
            else:
                cell.alignment = _align("center" if ci in (2, 3) else "left")
        rn += 1

    ws.merge_cells(f"A{rn}:C{rn}")
    ws[f"A{rn}"].value = "TOTAL"
    ws[f"A{rn}"].font = _font(bold=True, color=C_WHITE, size=9)
    ws[f"A{rn}"].fill = _fill(C_TEAL)
    ws[f"A{rn}"].alignment = _align("right")
    ws[f"A{rn}"].border = _border()
    ws[f"D{rn}"].value = f"=SUM(D4:D{rn-1})"
    ws[f"D{rn}"].font = _font(bold=True, color=C_WHITE, size=9)
    ws[f"D{rn}"].fill = _fill(C_TEAL)
    ws[f"D{rn}"].number_format = 'R$ #,##0.00'
    ws[f"D{rn}"].alignment = _align("right")
    ws[f"D{rn}"].border = _border()
    ws[f"E{rn}"].value = "100,0%"
    ws[f"E{rn}"].font = _font(bold=True, color=C_WHITE, size=9)
    ws[f"E{rn}"].fill = _fill(C_TEAL)
    ws[f"E{rn}"].alignment = _align("center")
    ws[f"E{rn}"].border = _border()

    note = rn + 2
    ws.merge_cells(f"A{note}:E{note}")
    ws[f"A{note}"].value = "* 'Total quantidade' omitido — unidades distintas (m², m, un) não são somáveis."
    ws[f"A{note}"].font = _font(italic=True, size=8, color="777777")
    ws[f"A{note}"].alignment = _align("left", wrap=True)
    ws.row_dimensions[note].height = 20


def _aba_criterios(wb):
    ws = wb.create_sheet("Critérios Adotados")

    ws.merge_cells("A1:D1")
    ws["A1"] = "Critérios de medição e referências normativas"
    ws["A1"].font = _font(bold=True, color=C_WHITE, size=12)
    ws["A1"].fill = _fill(C_BLUE_DARK)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 20

    # Seção critérios
    ws.merge_cells("A3:D3")
    ws["A3"].value = "Critérios de medição por elemento"
    ws["A3"].font = _font(bold=True, color=C_WHITE, size=10)
    ws["A3"].fill = _fill(C_BLUE_MED)
    ws["A3"].alignment = _align("center")
    ws["A3"].border = _border()
    for col in ("B", "C", "D"):
        ws[f"{col}3"].fill = _fill(C_BLUE_MED)
        ws[f"{col}3"].border = _border()

    for ci, (h, w) in enumerate(zip(
        ["Elemento", "Entidade IFC", "Unidade", "Critério de medição adotado"],
        [25, 22, 9, 60]
    ), 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = _font(bold=True, size=9)
        cell.fill = _fill(C_GRAY_HEAD)
        cell.alignment = _align("center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    for i, (elem, ifc, un, crit) in enumerate(CRITERIOS, 5):
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
        for ci, val in enumerate([elem, ifc, un, crit], 1):
            cell = ws.cell(row=i, column=ci, value=val)
            cell.font = _font(size=9)
            cell.fill = _fill(bg)
            cell.alignment = _align("left", v="top", wrap=(ci == 4))
            cell.border = _border()
        ws.row_dimensions[i].height = 28

    # Seção normas
    ns = len(CRITERIOS) + 7
    ws.merge_cells(f"A{ns}:D{ns}")
    ws[f"A{ns}"].value = "Referências normativas e de mercado"
    ws[f"A{ns}"].font = _font(bold=True, color=C_WHITE, size=10)
    ws[f"A{ns}"].fill = _fill(C_BLUE_MED)
    ws[f"A{ns}"].alignment = _align("center")
    ws[f"A{ns}"].border = _border()
    for col in ("B", "C", "D"):
        ws[f"{col}{ns}"].fill = _fill(C_BLUE_MED)
        ws[f"{col}{ns}"].border = _border()

    for i, (cod, desc) in enumerate(NORMAS):
        rn = ns + 1 + i
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
        ws.merge_cells(f"B{rn}:D{rn}")
        for cl, val in zip(("A", "B"), [cod, desc]):
            cell = ws[f"{cl}{rn}"]
            cell.value = val
            cell.font = _font(bold=(cl == "A"), size=9)
            cell.fill = _fill(bg)
            cell.alignment = _align("left", wrap=True)
            cell.border = _border()
        for col in ("C", "D"):
            ws[f"{col}{rn}"].fill = _fill(bg)
            ws[f"{col}{rn}"].border = _border()
        ws.row_dimensions[rn].height = 22

    # Seção mapeamento IFC
    ifc_start = ns + len(NORMAS) + 3
    ws.merge_cells(f"A{ifc_start}:D{ifc_start}")
    ws[f"A{ifc_start}"].value = "Mapeamento de entidades IFC utilizadas"
    ws[f"A{ifc_start}"].font = _font(bold=True, color=C_WHITE, size=10)
    ws[f"A{ifc_start}"].fill = _fill(C_BLUE_MED)
    ws[f"A{ifc_start}"].alignment = _align("center")
    ws[f"A{ifc_start}"].border = _border()
    for col in ("B", "C", "D"):
        ws[f"{col}{ifc_start}"].fill = _fill(C_BLUE_MED)
        ws[f"{col}{ifc_start}"].border = _border()

    heads_row = ifc_start + 1
    for ci, h in enumerate(["Entidade IFC", "PredefinedType", "Uso no QuantAI"], 1):
        cell = ws.cell(row=heads_row, column=ci, value=h)
        cell.font = _font(bold=True, size=9)
        cell.fill = _fill(C_GRAY_HEAD)
        cell.alignment = _align("center")
        cell.border = _border()
    ws.cell(row=heads_row, column=4).fill = _fill(C_GRAY_HEAD)
    ws.cell(row=heads_row, column=4).border = _border()

    for i, (ent, pt, uso) in enumerate(IFC_MAP):
        rn = heads_row + 1 + i
        bg = C_BLUE_LIGHT if i % 2 == 0 else C_WHITE
        ws.merge_cells(f"C{rn}:D{rn}")
        for ci, val in enumerate([ent, pt, uso], 1):
            cell = ws.cell(row=rn, column=ci, value=val)
            cell.font = _font(size=9, bold=(ci == 1))
            cell.fill = _fill(bg)
            cell.alignment = _align("left")
            cell.border = _border()
        ws.cell(row=rn, column=4).fill = _fill(bg)
        ws.cell(row=rn, column=4).border = _border()


def gerar_excel_bytes(rows: list[dict], estado: str, regime: str) -> bytes:
    """
    Gera o workbook em memória e retorna os bytes para download via Streamlit.

    Parâmetros
    ----------
    rows   : lista de dicts (saída de extrair_quantitativos + preços da API)
    estado : sigla do estado (ex: 'SP')
    regime : 'desonerado' ou 'nao_desonerado'

    Retorna
    -------
    bytes do arquivo .xlsx prontos para st.download_button
    """
    wb = Workbook()
    _aba_quantitativos(wb, rows, estado, regime)
    _aba_resumo(wb, rows)
    _aba_criterios(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
